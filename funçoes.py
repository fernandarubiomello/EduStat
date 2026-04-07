# FUNÇÕES:

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
# p tirar esse aviso: C:\Users\55169\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.11_qbz5n2kfra8p0\LocalCache\local-packages\Python311\site-packages\openpyxl\worksheet\_reader.py:329: UserWarning: Unknown extension is not supported and will be removed warn(msg) e outros mais que poderiam aparecer

# import openpyxl
# import csv
# import os
# import statistics as st
# from openpyxl.styles import PatternFill
# from openpyxl.drawing.image import Image
# import matplotlib.pyplot as plt

import csv
import os
import statistics as st
import sys


# verificando se a pessoa tem a bibliotecas openpyxl e matplotlib instaladas para poder importá-las
try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.drawing.image import Image
except ModuleNotFoundError:
    print("\n✘ O módulo 'openpyxl' não está instalado.")
    print("→ Digite no terminal: pip install openpyxl\n")
    exit()

try:
    import matplotlib.pyplot as plt
except ModuleNotFoundError:
    print("\n✘ O módulo 'matplotlib' não está instalado.")
    print("→ Digite no terminal: pip install matplotlib\n")
    exit()

from dados import *

# variaveis globais
NUMERO_LINHAS = 0
NUMERO_COLUNAS = 0
TODAS_MEDIAS = []

VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # acima da média
VERMELHO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # abaixo da média

#caminho CSV
def identificarCaminhoCSV(arquivos_csv, n, dados):

    caminho = arquivos_csv[n-1]
    print(f"✔ Arquivo CSV encontrado: {os.path.basename(caminho)}")
    
    validar = validarCSV(caminho)

    if (validar):
        criarDicCSV(dados, caminho)
    else:
        print("\n Agradeçemos o uso do EduStat ★")
        print("→ Leia as instruções e tente novamente!")

    return

#caminho XLS
def identificarCaminhoXLS(arquivos_xls, n, dados):

    caminho = arquivos_xls[n-1]
    print(f"✔ Arquivo Excel encontrado: {os.path.basename(caminho)}")

    validar = validarPlanilha(caminho)

    if (validar):
        criarDicXLS(dados, caminho)
    else:
        print("\n Agradeçemos o uso do EduStat ★")
        print("→ Leia as instruções e tente novamente!")

    return

# verificando se o csv veio correto
def validarCSV(caminho):

    erros = []

    try:
        with open(caminho, 'r', encoding='utf-8') as arquivo:
            leitor = list(csv.reader(arquivo, delimiter=','))

    except UnicodeDecodeError:
        print("✘ ERRO DE CODIFICAÇÃO NO CSV:")
        print("O arquivo não está salvo em UTF-8.")
        print("""→ Como corrigir:
1. Abra o arquivo CSV no Excel / Google Sheets / LibreOffice.
2. Clique em “Salvar como…”.
3. Selecione “CSV UTF-8 (separado por vírgulas)”.
4. Salve e rode novamente o EduStat.""")
        return False

    linhasLimpas = [] # conteudo do csv sem a as linhas vazias

    for linha in leitor:

        if not any(c.strip() for c in linha):  # pelo menos uma célula não vazia
            continue
        linhasLimpas.append(linha)

    if len(linhasLimpas) < 2:
        erros.append("Arquivo CSV vazio ou sem dados após o cabeçalho.")
        return False

    primeiraLinha = linhasLimpas[0]

    # tirando os vazios do final
    while len(primeiraLinha) > 1 and primeiraLinha[-1].strip() == "":
        primeiraLinha.pop()

    # print(f"-> {primeiraLinha}")
    # ex: -> ['\ufeff', 'Nota_Prova1', 'Nota_Prova2', 'Nota_Trabalho']

    # cabeçalho
    if primeiraLinha[0].strip() == "" or primeiraLinha[0].strip() == "\ufeff":
        erros.append("Célula A1 deve conter o nome da coluna de matérias, deve ser texto (ex: 'Matéria').")

    for colIdx, valor in enumerate(primeiraLinha[1:], start = 2):

        if valor.strip() == "" or valor.strip() == "\ufeff":
            erros.append(f"Célula da coluna {colIdx} na primeira linha deve ser texto (ex: Nota 1).")

    # linhas
    for linIdx, linha in enumerate(linhasLimpas[1:], start = 2):

        while len(linha) > 1 and linha[-1].strip() == "": # removendo colunas vazias nas linhas
            linha.pop()

        if len(linha) == 0:
            continue

        materia = linha[0]

        if materia.strip() == "" or materia.strip() == "\ufeff":
            erros.append(f"Valor inválido em A{linIdx}: o nome da matéria deve ser texto.")
            continue

        # se só tem matéria e mais nd
        if len(linha) == 1:
            erros.append(f"A matéria em A{linIdx} está sem notas.")
            continue

        # notas
        for colIdx, valor in enumerate(linha[1:], start=2):

            letraColuna = openpyxl.utils.get_column_letter(colIdx)

            if valor.strip() == "" or valor.strip() == "\ufeff":
                erros.append(f"Valor inválido em {letraColuna}{linIdx}: notas vazias não são permitidas no meio dos dados.")
                continue

            try:
                float(valor.replace(",", "."))

            except:
                erros.append(f"Valor inválido em {letraColuna}{linIdx}: deve ser um número (ex: 7.5).")

    if erros:
        print("✘ ERROS ENCONTRADOS NO CSV:")
        for e in erros:
            print(" -", e)
        return False

    print("✔ CSV validado com sucesso!")
    return True


# verificando se a planilha veio correta
def validarPlanilha(caminho):

    try:
        wb = openpyxl.load_workbook(caminho)
    except Exception as erro:
        print("\n ✘ ERRO AO LER A PLANILHA:")
        print(f"Erro: {str(erro)}")
        print("Nosso sistema encontrou formatações avançadas ou invisíveis na sua planilha. Limpe os estilos da planilha e/ou tente limpar as células adjacentes as principais, salvar a planilha e mandar ela novamente para o EduStat.")
        exit()

    ws = wb.active

    # excluir linhas vazias:
    linhasApagar = []

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if all(cel is None for cel in row):
            linhasApagar.append(idx)

    # ta da ultima p primeira, p nao ter problema com indice
    for linha in reversed(linhasApagar):
        ws.delete_rows(linha)

    wb.save(caminho)

    erros = []

    # parte do cabecalho
    primeiraLinha = list(ws.iter_rows(min_row = 1, max_row = 1, values_only = True))[0]

    # p tirar os None dos finais
    ultimoNone = 0
    for i in range(len(primeiraLinha)):
        if (primeiraLinha[i] != None):
            ultimoNone = i
    primeiraLinha = primeiraLinha[:ultimoNone+1]

    if primeiraLinha[0] is None or not isinstance(primeiraLinha[0], str):

        erros.append("Célula A1 deve conter o nome da coluna de matérias, deve ser texto (ex: 'Matéria').")

    for col, valor in enumerate(primeiraLinha[1:], start = 2):

        if not isinstance(valor, str) or valor is None or str(valor).strip() == "":
            erros.append(f"Célula da coluna {col} na primeira linha deve ser texto (ex: Nota 1, Nota Prova, Avaliação).")

    # parte da coluna A e das notas
    for linhaIdx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # p tirar os None dos finais
        ultimoNone = 0
        listaRow = list(row)
        for i in range(len(listaRow)):
            if listaRow[i] is not None:
                ultimoNone = i
        listaRow = listaRow[:ultimoNone + 1]

        if len(listaRow) == 1: #quando so tem a materia
            celula = f"A{linhaIdx}"
            erros.append(f"A matéria em {celula} está sem notas.")
            continue

        # nome da materia
        if listaRow[0] is None or not isinstance(listaRow[0], str) or listaRow[0].strip() == "":
            celula = f"A{linhaIdx}"
            erros.append(f"Valor inválido em {celula}: o nome da matéria deve ser texto.")

        # notas
        for colIdx, valor in enumerate(listaRow[1:], start=2): 
            letraColuna = openpyxl.utils.get_column_letter(colIdx)
            celula = f"{letraColuna}{linhaIdx}"

            if isinstance(valor, float) or isinstance(valor, int):
                continue

            erros.append(f"Valor inválido em {celula}: deve ser um número (ex: 7.5).")


    if erros:

        print("✘ ERROS NA PLANILHA ENCONTRADOS:")

        for e in erros:
            print(" -", e)
        return False

    print("✔ Planilha validada com sucesso!")

    return True

# criando o dicionário com os dados do arquivo csv
def criarDicCSV(dados, caminho):

    global NUMERO_COLUNAS, NUMERO_LINHAS # assim mostra que essas var vêm de fora da função e vão ser modificadas aqui

    with open(caminho, 'r', encoding='utf-8') as arquivo:

        leitor_csv = csv.reader(arquivo, delimiter=',')

        linhasCSV = list(leitor_csv)

        for linha in linhasCSV[1:]: #pulando a 1º linha (tipo com next(leitor_csv))

            quantNotas = len(linha) - 1
            nome = linha[0]

            if  str(nome).strip() == "":
                continue

            dados[nome] = []

            for i in range(1, quantNotas+1):
                if (linha[i].strip() != ''):
                    valor = linha[i].replace(",", ".")
                    dados[nome].append(float(valor))

        # criando um xlsm com os dados do csv
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notas"
        for linha in linhasCSV:
            if (not any(c.strip() for c in linha)): # nao coloca as vazias (verificacao de novo)
                continue
            ws.append(linha)
        novoCaminho = f"{os.path.splitext(caminho)[0]}.xlsx"
        wb.save(novoCaminho)

        # pegando número de linhas e de colunas que a planilha possui, antes dela ser editada
        #NUMERO_LINHAS = len([celula for celula in ws['A'] if celula.value is not None]) # conta A1, A2, A3 ..., com valores != nulo, assim, conta quantas linhas a planilha tem
        NUMERO_LINHAS = len(dados)+1
        #NUMERO_COLUNAS = len([c for c in ws[1] if c.value is not None]) # conta A1, B1, C1, ..., com valores != nulo, assim, contas quantas colunas a planilha tem
        maxColNotas = max(len(v) for v in dados.values())
        NUMERO_COLUNAS = maxColNotas + 1

        modificarPlanilha(dados, novoCaminho)

        return

#criando o dicionário com os dados da planilha
def criarDicXLS(dados, caminho):

    global NUMERO_COLUNAS, NUMERO_LINHAS # assim mostra que essas var vêm de fora da função e vão ser modificadas aqui

    try:
        wb = openpyxl.load_workbook(caminho)
    except Exception as erro:
        print("\n ✘ ERRO AO LER A PLANILHA:")
        print(f"Erro: {str(erro)}")
        print("Nosso sistema encontrou formatações avançadas ou invisíveis na sua planilha. Limpe os estilos da planilha e/ou tente limpar as células adjacentes as principais, salvar a planilha e mandar ela novamente para o EduStat.")
        exit()

    ws = wb.active

    for row in ws.iter_rows(min_row = 2, max_row = ws.max_row, values_only = True):

        row = list(row)
        nome = row[0]

        if nome is None or str(nome).strip() == "":
            continue

        del row[0]

        notas = []
        for i in row:
            if i is not None:
                notas.append(float(i))

        dados[nome] = notas

    wb.save(caminho)

    # pegando número de linhas e de colunas que a planilha possui, antes dela ser editada
    #NUMERO_LINHAS = len([celula for celula in ws['A'] if celula.value is not None]) # conta A1, A2, A3 ..., com valores != nulo, assim, conta quantas linhas a planilha tem
    NUMERO_LINHAS = len(dados)+1
    #NUMERO_COLUNAS = len([c for c in ws[1] if c.value is not None]) # conta A1, B1, C1, ..., com valores != nulo, assim, contas quantas colunas a planilha tem
    maxColNotas = max(len(v) for v in dados.values())
    NUMERO_COLUNAS = maxColNotas + 1

    modificarPlanilha(dados, caminho)

    return

def modificarPlanilha(dados, caminho):

    try:
        wb = openpyxl.load_workbook(caminho)
    except Exception as erro:
        print("\n ✘ ERRO AO LER A PLANILHA:")
        print(f"Erro: {str(erro)}")
        print("Nosso sistema encontrou formatações avançadas ou invisíveis na sua planilha. Limpe os estilos da planilha e/ou tente limpar as células adjacentes as principais, salvar a planilha e mandar ela novamente para o EduStat.")
        exit()

    ws = wb.active

    inserirMediaDesvioVarianciaMateria(dados, ws)
    inserirMediaDesvioVarianciaTotal(dados, ws)
    inserirQuantoFaltaPassar(ws)
    inserirHistograma(ws)
    inserirBoxPlot(ws)

    wb.save(caminho)

    print("\nAgradeçemos o uso do EduStat ★ \n")
    print("Abra novamente sua planilha - ou sua nova planilha caso tenha mandando um csv - e veja suas estatísticas.")
    print("Até mais...")

    return

#inserindo média, desvio padrão e variancia de cada matéria na planilha
def inserirMediaDesvioVarianciaMateria(dados, ws):

    global TODAS_MEDIAS

    ws.cell(row = 1, column = NUMERO_COLUNAS + 1, value = "Média")
    ws.cell(row = 1, column = NUMERO_COLUNAS + 2, value = "Desvio Padrão")
    ws.cell(row = 1, column = NUMERO_COLUNAS + 3, value = "Variância")

    # pegando valores pela planilha (maior chance de erro)
    # for i in range(2, NUMERO_LINHAS+1):
        
    #     notas = []

    #     for j in range(2, NUMERO_COLUNAS+1):

    #         notas.append( float(ws.cell(row=i, column=j).value ))

    # pegando valores pelo dicionário
    for i, valor in enumerate(dados.values(), start = 2):

        # i -> indice,
        # valor -> valor de cada chave do dicionario
        notas = list(valor)  # copia a lista de notas

        if (len(notas) > 1):
            media = st.mean(notas)
            desvioPadrao = st.stdev(notas)
            variancia = st.variance(notas)
        else:
            media = notas[0]
            desvioPadrao = 0
            variancia = 0

        TODAS_MEDIAS.append(media)

        celula = ws.cell(row = i, column = NUMERO_COLUNAS+1, value = media)
        if (media >= 5):
            celula.fill = VERDE
        else:
            celula.fill = VERMELHO
        celula.number_format = "0.00"

        ws.cell(row = i, column = NUMERO_COLUNAS+2, value = desvioPadrao).number_format = "0.00"
        ws.cell(row = i, column = NUMERO_COLUNAS+3, value = variancia).number_format = "0.00"

    print("✔ Média, Desvio Padrão e Variância por matéria adicionados na planilha com sucesso!")

    return

#inserindo média, desvio padrão e variancia total na planilha
def inserirMediaDesvioVarianciaTotal(dados, ws):

    ws.cell(row = NUMERO_LINHAS+1, column = 1, value = "Geral:")

    todasNotas = []

    for valor in dados.values():

        for i in list(valor):
            todasNotas.append(i)

    if (len(todasNotas) > 1):
        media = st.mean(todasNotas)
        desvioPadrao = st.pstdev(todasNotas)
        variancia = st.pvariance(todasNotas)
    else:
        media = todasNotas[0]
        desvioPadrao = 0
        variancia = 0

    celula = ws.cell(row = NUMERO_LINHAS+1, column = NUMERO_COLUNAS+1, value = media)
    if (media >= 5):
        celula.fill = VERDE
    else:
        celula.fill = VERMELHO
    celula.number_format = "0.00"

    ws.cell(row = NUMERO_LINHAS+1, column = NUMERO_COLUNAS+2, value = desvioPadrao).number_format = "0.00"
    ws.cell(row = NUMERO_LINHAS+1, column = NUMERO_COLUNAS+3, value = variancia).number_format = "0.00"

    print("✔ Média, Desvio Padrão e Variância total adicionados na planilha com sucesso!")

    return

# inserindo quanto de nota falta o aluno passar na matéria (média 5, como na USP)
def inserirQuantoFaltaPassar(ws):

    ws.cell(row = 1, column = NUMERO_COLUNAS+4, value = "Situação")

    for i, media in enumerate(TODAS_MEDIAS):

        falta = 10 - media

        if media > 5:
            ws.cell(row = i+2, column = NUMERO_COLUNAS+4, value = "Aprovado")

        elif media > 3:
            ws.cell(row = i+2, column = NUMERO_COLUNAS+4, value = f"Recuperação: falta {falta}")

        else:
            ws.cell(row = i+2, column = NUMERO_COLUNAS+4, value = f"Reprovado: faltou {falta}")

    print("✔ O quanto falta para passar adicionado na planilha com sucesso!")

    return

# criando o histograma com as médias das notas de cada matéria e inserindo ele na planilha
def inserirHistograma(ws):

    plt.hist(
        TODAS_MEDIAS,
        bins = 5,
        color = "#4C72B0",   
        edgecolor = "black", 
    )

    plt.title("Histograma: Distribuição das Médias por Matéria", fontsize = 14)
    plt.xlabel("Média das Notas", fontsize = 12)
    plt.ylabel("Frequência", fontsize = 12)

    plt.tight_layout() # ajusta automaticamento os entresubespaços

    plt.savefig("dados/histograma.png", dpi = 300)

    plt.close()

    img = Image("dados/histograma.png")
    img.width = 350
    img.height = 250
    ws.add_image(img, f"A{NUMERO_LINHAS+3}")

    print("✔ Histograma adicionado na planilha com sucesso! Sua imagem também pode ser encontrada na pasta 'dados'. Aproveite!")

    return

# criando o boxplot com as médias das notas de cada matéria e inserindo ele na planilha
def inserirBoxPlot(ws):
    
    plt.boxplot(
        TODAS_MEDIAS,
        patch_artist = True,  
        boxprops = dict(facecolor="#4C72B0", color="black"), 
        medianprops = dict(color="red", linewidth=2), # linha da mediana
        flierprops=dict(
            marker="o", # circulo
            markerfacecolor="red",
            markersize=8
        )  # outliers
    )

    plt.title("Boxplot: Distribuição das Médias por Matéria")
    plt.ylabel("Valores das Médias", fontsize = 12)

    plt.tight_layout() # ajusta automaticamento os entresubespaços

    plt.savefig("dados/boxplot.png", dpi = 300)

    plt.close()

    img = Image("dados/boxplot.png")
    img.width = 350
    img.height = 250
    ws.add_image(img, f"J{NUMERO_LINHAS+3}")

    print("✔ Boxplot adicionado na planilha com sucesso! Sua imagem também pode ser encontrada na pasta 'dados'. Aproveite!")

    return
